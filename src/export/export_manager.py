# -*- coding: utf-8 -*-
"""Export Manager - PDF and Excel export functionality"""
from datetime import datetime
from io import BytesIO
from typing import List, Tuple, Optional
import logging

from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExportManager:
    """Manager for exporting calculations to PDF and Excel formats"""
    
    # Formatting constants
    CURRENCY_FORMAT = "R$ {:.2f}"
    PERCENTAGE_FORMAT = "{:.1f}%"
    HEADER_COLOR = colors.HexColor("#1f77b4")
    HEADER_TEXT_COLOR = colors.whitesmoke
    ALTERNATE_ROW_COLOR = colors.HexColor("#f0f0f0")
    
    @staticmethod
    def export_to_pdf(calculations: List[dict], filename: Optional[str] = None) -> Tuple[bool, Optional[BytesIO], Optional[str]]:
        """
        Export calculations to PDF format
        
        Args:
            calculations: List of calculation dictionaries
            filename: Optional custom filename
            
        Returns:
            Tuple[success, pdf_buffer, error_message]
        """
        try:
            if not calculations:
                return False, None, "Nenhum cálculo para exportar"
            
            # Create PDF buffer
            pdf_buffer = BytesIO()
            doc = SimpleDocTemplate(
                pdf_buffer,
                pagesize=A4,
                rightMargin=0.75 * inch,
                leftMargin=0.75 * inch,
                topMargin=0.75 * inch,
                bottomMargin=0.75 * inch,
            )
            
            elements = []
            styles = getSampleStyleSheet()
            
            # Custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=ExportManager.HEADER_COLOR,
                spaceAfter=30,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=ExportManager.HEADER_COLOR,
                spaceAfter=12,
                spaceBefore=12,
                fontName='Helvetica-Bold'
            )
            
            # Title
            elements.append(Paragraph("Relatório de Cálculos de ROI em RPA", title_style))
            elements.append(Spacer(1, 0.3 * inch))
            
            # Date
            date_text = f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
            elements.append(Paragraph(date_text, styles['Normal']))
            elements.append(Spacer(1, 0.3 * inch))
            
            # Process each calculation
            for idx, calc in enumerate(calculations, 1):
                if idx > 1:
                    elements.append(PageBreak())
                
                # Process header
                process_name = calc.get('process_name', 'Sem Nome')
                elements.append(Paragraph(f"{idx}. {process_name}", heading_style))
                
                # Basic information
                basic_info = [
                    ['Campo', 'Valor'],
                    ['Departamento', calc.get('department', '—')],
                    ['Complexidade', calc.get('complexity', '—')],
                    ['Pessoas Envolvidas', str(calc.get('people_involved', 0))],
                    ['Sistemas', str(calc.get('systems_quantity', 0))],
                    ['Transações Diárias', f"{calc.get('daily_transactions', 0):,.0f}"],
                ]
                
                basic_table = Table(basic_info, colWidths=[2.5 * inch, 3.5 * inch])
                basic_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), ExportManager.HEADER_COLOR),
                    ('TEXTCOLOR', (0, 0), (-1, 0), ExportManager.HEADER_TEXT_COLOR),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, ExportManager.ALTERNATE_ROW_COLOR]),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                ]))
                
                elements.append(basic_table)
                elements.append(Spacer(1, 0.2 * inch))
                
                # Financial information
                elements.append(Paragraph("Informações Financeiras", ParagraphStyle(
                    'SubHeading',
                    parent=styles['Heading3'],
                    fontSize=12,
                    textColor=ExportManager.HEADER_COLOR,
                    spaceAfter=10,
                )))
                
                hourly_rate = calc.get('hourly_rate', 0)
                current_time = calc.get('current_time_per_month', 0)
                current_cost = hourly_rate * current_time
                
                financial_info = [
                    ['Métrica', 'Valor'],
                    ['Valor hora', ExportManager.CURRENCY_FORMAT.format(hourly_rate)],
                    ['Tempo Atual (horas/mês)', f"{current_time:.1f}"],
                    ['Custo Atual (mensal)', ExportManager.CURRENCY_FORMAT.format(current_cost)],
                    ['Custo Implementação RPA', ExportManager.CURRENCY_FORMAT.format(calc.get('rpa_implementation_cost', 0))],
                    ['Custo Mensal RPA', ExportManager.CURRENCY_FORMAT.format(calc.get('rpa_monthly_cost', 0))],
                    ['Custo Manutenção (%)', ExportManager.PERCENTAGE_FORMAT.format(calc.get('maintenance_percentage', 0))],
                    ['Custo Infra/Licença (mensal)', ExportManager.CURRENCY_FORMAT.format(calc.get('infra_license_cost', 0))],
                    ['Outros Custos (mensal)', ExportManager.CURRENCY_FORMAT.format(calc.get('other_costs', 0))],
                ]
                
                financial_table = Table(financial_info, colWidths=[2.5 * inch, 3.5 * inch])
                financial_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), ExportManager.HEADER_COLOR),
                    ('TEXTCOLOR', (0, 0), (-1, 0), ExportManager.HEADER_TEXT_COLOR),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, ExportManager.ALTERNATE_ROW_COLOR]),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                ]))
                
                elements.append(financial_table)
                elements.append(Spacer(1, 0.2 * inch))
                
                # ROI Results
                elements.append(Paragraph("Resultados de ROI", ParagraphStyle(
                    'SubHeading',
                    parent=styles['Heading3'],
                    fontSize=12,
                    textColor=ExportManager.HEADER_COLOR,
                    spaceAfter=10,
                )))
                
                roi_results = [
                    ['Período', 'Economia', 'ROI', 'ROI (%)'],
                    ['1 Ano', 
                     ExportManager.CURRENCY_FORMAT.format(calc.get('roi_first_year', 0)),
                     ExportManager.CURRENCY_FORMAT.format(calc.get('roi_first_year', 0)),
                     ExportManager.PERCENTAGE_FORMAT.format(calc.get('roi_percentage_first_year', 0))],
                    ['Payback (meses)', f"{calc.get('payback_period_months', 0):.1f}", '—', '—'],
                    ['Economia Mensal', ExportManager.CURRENCY_FORMAT.format(calc.get('monthly_savings', 0)), '—', '—'],
                    ['Economia Anual', ExportManager.CURRENCY_FORMAT.format(calc.get('annual_savings', 0)), '—', '—'],
                ]
                
                roi_table = Table(roi_results, colWidths=[1.5 * inch, 1.75 * inch, 1.75 * inch, 1.5 * inch])
                roi_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), ExportManager.HEADER_COLOR),
                    ('TEXTCOLOR', (0, 0), (-1, 0), ExportManager.HEADER_TEXT_COLOR),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, ExportManager.ALTERNATE_ROW_COLOR]),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                ]))
                
                elements.append(roi_table)
                elements.append(Spacer(1, 0.3 * inch))
            
            # Build PDF
            doc.build(elements)
            pdf_buffer.seek(0)
            
            return True, pdf_buffer, None
            
        except Exception as e:
            logger.error(f"Erro ao exportar PDF: {str(e)}")
            return False, None, f"Erro ao gerar PDF: {str(e)}"
    
    @staticmethod
    def export_to_excel(calculations: List[dict], filename: Optional[str] = None) -> Tuple[bool, Optional[BytesIO], Optional[str]]:
        """
        Export calculations to Excel format
        
        Args:
            calculations: List of calculation dictionaries
            filename: Optional custom filename
            
        Returns:
            Tuple[success, excel_buffer, error_message]
        """
        try:
            if not calculations:
                return False, None, "Nenhum cálculo para exportar"
            
            wb = Workbook()
            
            # Remove default sheet and create summary sheet
            if wb.active is not None:
                wb.remove(wb.active)
            summary_sheet = wb.create_sheet("Resumo", 0)
            
            # Configure summary sheet
            summary_sheet.column_dimensions['A'].width = 25
            summary_sheet.column_dimensions['B'].width = 15
            summary_sheet.column_dimensions['C'].width = 15
            summary_sheet.column_dimensions['D'].width = 15
            summary_sheet.column_dimensions['E'].width = 15
            
            # Header style
            header_fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Summary headers
            headers = ['Processo', 'Economia Mensal (R$)', 'Economia Anual (R$)', 'ROI 1º Ano (R$)', 'ROI %']
            for col, header in enumerate(headers, 1):
                cell = summary_sheet.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            # Summary data
            row = 2
            for calc in calculations:
                summary_sheet.cell(row=row, column=1).value = calc.get('process_name', '—')
                summary_sheet.cell(row=row, column=2).value = calc.get('monthly_savings', 0)
                summary_sheet.cell(row=row, column=3).value = calc.get('annual_savings', 0)
                summary_sheet.cell(row=row, column=4).value = calc.get('roi_first_year', 0)
                summary_sheet.cell(row=row, column=5).value = calc.get('roi_percentage_first_year', 0)
                
                # Format currency and percentage
                summary_sheet.cell(row=row, column=2).number_format = '"R$" #,##0.00'
                summary_sheet.cell(row=row, column=3).number_format = '"R$" #,##0.00'
                summary_sheet.cell(row=row, column=4).number_format = '"R$" #,##0.00'
                summary_sheet.cell(row=row, column=5).number_format = '0.0"%"'
                
                # Borders and alternating background
                if row % 2 == 0:
                    alt_fill = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")
                    for col in range(1, 6):
                        summary_sheet.cell(row=row, column=col).fill = alt_fill
                
                for col in range(1, 6):
                    summary_sheet.cell(row=row, column=col).border = border
                
                row += 1
            
            # Create detailed sheets for each calculation
            for calc in calculations:
                process_name = calc.get('process_name', 'Processo')
                # Limit sheet name to 31 characters
                sheet_name = process_name[:31] if len(process_name) <= 31 else process_name[:28] + "..."
                
                # Avoid duplicate sheet names
                sheet_counter = 1
                original_name = sheet_name
                while sheet_name in wb.sheetnames:
                    sheet_name = f"{original_name[:20]}_{sheet_counter}"
                    sheet_counter += 1
                
                detail_sheet = wb.create_sheet(sheet_name)
                
                # Configure columns
                for col in range(1, 3):
                    detail_sheet.column_dimensions[get_column_letter(col)].width = 25 if col == 1 else 20
                
                # Write process details
                row = 1
                
                # Title
                detail_sheet.cell(row=row, column=1).value = f"Detalhes: {process_name}"
                title_cell = detail_sheet.cell(row=row, column=1)
                title_cell.font = Font(bold=True, size=14, color="1f77b4")
                row += 2
                
                # Basic information
                detail_sheet.cell(row=row, column=1).value = "INFORMAÇÕES BÁSICAS"
                detail_sheet.cell(row=row, column=1).font = Font(bold=True, size=11)
                row += 1
                
                basic_fields = [
                    ('Departamento', calc.get('department', '—')),
                    ('Complexidade', calc.get('complexity', '—')),
                    ('Pessoas Envolvidas', calc.get('people_involved', 0)),
                    ('Sistemas', calc.get('systems_quantity', 0)),
                    ('Transações Diárias', calc.get('daily_transactions', 0)),
                ]
                
                for label, value in basic_fields:
                    detail_sheet.cell(row=row, column=1).value = label
                    detail_sheet.cell(row=row, column=2).value = value
                    detail_sheet.cell(row=row, column=1).font = Font(bold=True)
                    row += 1
                
                row += 1
                
                # Financial information
                detail_sheet.cell(row=row, column=1).value = "INFORMAÇÕES FINANCEIRAS"
                detail_sheet.cell(row=row, column=1).font = Font(bold=True, size=11)
                row += 1
                
                financial_fields = [
                    ('Valor hora (R$)', calc.get('hourly_rate', 0)),
                    ('Tempo Atual (horas/mês)', calc.get('current_time_per_month', 0)),
                    ('Custo Implementação (R$)', calc.get('rpa_implementation_cost', 0)),
                    ('Custo Mensal RPA (R$)', calc.get('rpa_monthly_cost', 0)),
                    ('Manutenção (%)', calc.get('maintenance_percentage', 0)),
                    ('Infra/Licença (R$/mês)', calc.get('infra_license_cost', 0)),
                    ('Outros Custos (R$/mês)', calc.get('other_costs', 0)),
                ]
                
                for label, value in financial_fields:
                    detail_sheet.cell(row=row, column=1).value = label
                    cell = detail_sheet.cell(row=row, column=2)
                    cell.value = value
                    if 'R$' in label:
                        cell.number_format = '"R$" #,##0.00'
                    elif '%' in label:
                        cell.number_format = '0.0"%"'
                    detail_sheet.cell(row=row, column=1).font = Font(bold=True)
                    row += 1
                
                row += 1
                
                # ROI Results
                detail_sheet.cell(row=row, column=1).value = "RESULTADOS DE ROI"
                detail_sheet.cell(row=row, column=1).font = Font(bold=True, size=11)
                row += 1
                
                roi_fields = [
                    ('Economia Mensal (R$)', calc.get('monthly_savings', 0)),
                    ('Economia Anual (R$)', calc.get('annual_savings', 0)),
                    ('Economia 1º Ano (R$)', calc.get('roi_first_year', 0)),
                    ('ROI 1º Ano (%)', calc.get('roi_percentage_first_year', 0)),
                    ('Payback (meses)', calc.get('payback_period_months', 0)),
                ]
                
                for label, value in roi_fields:
                    detail_sheet.cell(row=row, column=1).value = label
                    cell = detail_sheet.cell(row=row, column=2)
                    cell.value = value
                    if 'R$' in label:
                        cell.number_format = '"R$" #,##0.00'
                    elif '%' in label:
                        cell.number_format = '0.0"%"'
                    detail_sheet.cell(row=row, column=1).font = Font(bold=True)
                    row += 1
            
            # Save to buffer
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return True, excel_buffer, None
            
        except Exception as e:
            logger.error(f"Erro ao exportar Excel: {str(e)}")
            return False, None, f"Erro ao gerar Excel: {str(e)}"

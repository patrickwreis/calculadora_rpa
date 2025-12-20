# -*- coding: utf-8 -*-
"""Tests for Export Manager - PDF and Excel export functionality"""
import pytest
from io import BytesIO
from src.export import ExportManager


class TestPDFExport:
    """Test PDF export functionality"""
    
    @pytest.fixture
    def sample_calculations(self):
        """Sample calculation data for export"""
        return [
            {
                'process_name': 'Teste PDF 1',
                'department': 'TI',
                'complexity': 'Média',
                'people_involved': 2,
                'systems_quantity': 3,
                'daily_transactions': 1000,
                'hourly_rate': 50.0,
                'current_time_per_month': 160.0,
                'rpa_implementation_cost': 5000.0,
                'rpa_monthly_cost': 200.0,
                'maintenance_percentage': 10.0,
                'infra_license_cost': 500.0,
                'other_costs': 1000.0,
                'monthly_savings': 2500.0,
                'annual_savings': 30000.0,
                'roi_first_year': 25000.0,
                'roi_percentage_first_year': 500.0,
                'payback_period_months': 2.4,
            },
            {
                'process_name': 'Teste PDF 2',
                'department': 'Financeiro',
                'complexity': 'Alta',
                'people_involved': 3,
                'systems_quantity': 5,
                'daily_transactions': 2000,
                'hourly_rate': 75.0,
                'current_time_per_month': 200.0,
                'rpa_implementation_cost': 8000.0,
                'rpa_monthly_cost': 300.0,
                'maintenance_percentage': 15.0,
                'infra_license_cost': 600.0,
                'other_costs': 1500.0,
                'monthly_savings': 3500.0,
                'annual_savings': 42000.0,
                'roi_first_year': 34000.0,
                'roi_percentage_first_year': 425.0,
                'payback_period_months': 2.3,
            }
        ]
    
    def test_pdf_export_success(self, sample_calculations):
        """Test successful PDF export"""
        success, pdf_buffer, error_msg = ExportManager.export_to_pdf(sample_calculations)
        
        assert success is True
        assert error_msg is None
        assert pdf_buffer is not None
        assert isinstance(pdf_buffer, BytesIO)
        assert pdf_buffer.getbuffer().nbytes > 0
    
    def test_pdf_export_empty_list(self):
        """Test PDF export with empty calculation list"""
        success, pdf_buffer, error_msg = ExportManager.export_to_pdf([])
        
        assert success is False
        assert pdf_buffer is None
        assert error_msg is not None
        assert "nenhum cálculo" in error_msg.lower()
    
    def test_pdf_export_single_calculation(self, sample_calculations):
        """Test PDF export with single calculation"""
        success, pdf_buffer, error_msg = ExportManager.export_to_pdf([sample_calculations[0]])
        
        assert success is True
        assert error_msg is None
        assert pdf_buffer is not None
    
    def test_pdf_buffer_is_readable(self, sample_calculations):
        """Test that PDF buffer is properly seeked and readable"""
        success, pdf_buffer, error_msg = ExportManager.export_to_pdf(sample_calculations)
        
        assert success is True
        assert pdf_buffer is not None
        # Buffer should be at position 0
        assert pdf_buffer.tell() == 0
        # Should be able to read from it
        content = pdf_buffer.read(4)
        assert len(content) > 0
        # Reset position for reuse
        pdf_buffer.seek(0)


class TestExcelExport:
    """Test Excel export functionality"""
    
    @pytest.fixture
    def sample_calculations(self):
        """Sample calculation data for export"""
        return [
            {
                'process_name': 'Teste Excel 1',
                'department': 'Operações',
                'complexity': 'Baixa',
                'people_involved': 1,
                'systems_quantity': 2,
                'daily_transactions': 500,
                'hourly_rate': 40.0,
                'current_time_per_month': 100.0,
                'rpa_implementation_cost': 3000.0,
                'rpa_monthly_cost': 150.0,
                'maintenance_percentage': 8.0,
                'infra_license_cost': 400.0,
                'other_costs': 800.0,
                'monthly_savings': 1500.0,
                'annual_savings': 18000.0,
                'roi_first_year': 15000.0,
                'roi_percentage_first_year': 500.0,
                'payback_period_months': 2.0,
            },
            {
                'process_name': 'Teste Excel 2',
                'department': 'RH',
                'complexity': 'Média',
                'people_involved': 2,
                'systems_quantity': 4,
                'daily_transactions': 800,
                'hourly_rate': 60.0,
                'current_time_per_month': 150.0,
                'rpa_implementation_cost': 6000.0,
                'rpa_monthly_cost': 250.0,
                'maintenance_percentage': 12.0,
                'infra_license_cost': 450.0,
                'other_costs': 1200.0,
                'monthly_savings': 2200.0,
                'annual_savings': 26400.0,
                'roi_first_year': 20400.0,
                'roi_percentage_first_year': 340.0,
                'payback_period_months': 2.7,
            }
        ]
    
    def test_excel_export_success(self, sample_calculations):
        """Test successful Excel export"""
        success, excel_buffer, error_msg = ExportManager.export_to_excel(sample_calculations)
        
        assert success is True
        assert error_msg is None
        assert excel_buffer is not None
        assert isinstance(excel_buffer, BytesIO)
        assert excel_buffer.getbuffer().nbytes > 0
    
    def test_excel_export_empty_list(self):
        """Test Excel export with empty calculation list"""
        success, excel_buffer, error_msg = ExportManager.export_to_excel([])
        
        assert success is False
        assert excel_buffer is None
        assert error_msg is not None
        assert "nenhum cálculo" in error_msg.lower()
    
    def test_excel_export_single_calculation(self, sample_calculations):
        """Test Excel export with single calculation"""
        success, excel_buffer, error_msg = ExportManager.export_to_excel([sample_calculations[0]])
        
        assert success is True
        assert error_msg is None
        assert excel_buffer is not None
    
    def test_excel_buffer_is_readable(self, sample_calculations):
        """Test that Excel buffer is properly seeked and readable"""
        success, excel_buffer, error_msg = ExportManager.export_to_excel(sample_calculations)
        
        assert success is True
        assert excel_buffer is not None
        # Buffer should be at position 0
        assert excel_buffer.tell() == 0
        # Should be able to read from it
        content = excel_buffer.read(4)
        assert len(content) > 0
        # Reset position for reuse
        excel_buffer.seek(0)
    
    def test_excel_export_multiple_sheets(self, sample_calculations):
        """Test that Excel export creates multiple sheets"""
        success, excel_buffer, error_msg = ExportManager.export_to_excel(sample_calculations)
        
        assert success is True
        assert excel_buffer is not None
        excel_buffer.seek(0)
        
        # Load the workbook to verify sheets
        from openpyxl import load_workbook
        wb = load_workbook(excel_buffer)
        
        # Should have at least 3 sheets: Resumo + 2 detail sheets
        assert len(wb.sheetnames) >= 3
        assert 'Resumo' in wb.sheetnames
    
    def test_excel_export_handles_long_process_names(self):
        """Test Excel export handles long process names that exceed 31 char limit"""
        long_name = "A" * 50  # Excel sheet names limited to 31 chars
        calculations = [
            {
                'process_name': long_name,
                'department': 'Test',
                'complexity': 'Média',
                'people_involved': 1,
                'systems_quantity': 1,
                'daily_transactions': 100,
                'hourly_rate': 50.0,
                'current_time_per_month': 100.0,
                'rpa_implementation_cost': 5000.0,
                'rpa_monthly_cost': 200.0,
                'maintenance_percentage': 10.0,
                'infra_license_cost': 500.0,
                'other_costs': 1000.0,
                'monthly_savings': 2000.0,
                'annual_savings': 24000.0,
                'roi_first_year': 19000.0,
                'roi_percentage_first_year': 380.0,
                'payback_period_months': 3.0,
            }
        ]
        
        success, excel_buffer, error_msg = ExportManager.export_to_excel(calculations)
        
        assert success is True
        assert excel_buffer is not None
        excel_buffer.seek(0)
        
        from openpyxl import load_workbook
        wb = load_workbook(excel_buffer)
        
        # All sheet names should be valid (≤31 chars)
        for sheet_name in wb.sheetnames:
            assert len(sheet_name) <= 31


class TestExportIntegration:
    """Integration tests for export functionality"""
    
    @pytest.fixture
    def sample_calculations(self):
        """Sample calculation data for integration tests"""
        return [
            {
                'process_name': 'Integration Test Process',
                'department': 'IT',
                'complexity': 'Alta',
                'people_involved': 4,
                'systems_quantity': 8,
                'daily_transactions': 5000,
                'hourly_rate': 85.0,
                'current_time_per_month': 400.0,
                'rpa_implementation_cost': 15000.0,
                'rpa_monthly_cost': 500.0,
                'maintenance_percentage': 15.0,
                'infra_license_cost': 1000.0,
                'other_costs': 2000.0,
                'monthly_savings': 5000.0,
                'annual_savings': 60000.0,
                'roi_first_year': 45000.0,
                'roi_percentage_first_year': 300.0,
                'payback_period_months': 3.0,
            }
        ]
    
    def test_both_exports_succeed(self, sample_calculations):
        """Test that both PDF and Excel exports succeed"""
        pdf_success, pdf_buffer, pdf_error = ExportManager.export_to_pdf(sample_calculations)
        excel_success, excel_buffer, excel_error = ExportManager.export_to_excel(sample_calculations)
        
        assert pdf_success is True
        assert excel_success is True
        assert pdf_error is None
        assert excel_error is None
        assert pdf_buffer is not None
        assert excel_buffer is not None
    
    def test_exports_handle_missing_fields(self):
        """Test that exports handle missing optional fields gracefully"""
        incomplete_calc = {
            'process_name': 'Incomplete Process',
            # Minimal required fields only
        }
        
        pdf_success, _, pdf_error = ExportManager.export_to_pdf([incomplete_calc])
        excel_success, _, excel_error = ExportManager.export_to_excel([incomplete_calc])
        
        # Both should succeed even with incomplete data
        assert pdf_success is True
        assert excel_success is True
    
    def test_exports_with_zero_values(self):
        """Test exports handle zero financial values"""
        zero_calc = {
            'process_name': 'Zero Values Test',
            'department': 'Test',
            'complexity': 'Baixa',
            'people_involved': 0,
            'systems_quantity': 0,
            'daily_transactions': 0,
            'hourly_rate': 0.0,
            'current_time_per_month': 0.0,
            'rpa_implementation_cost': 0.0,
            'rpa_monthly_cost': 0.0,
            'maintenance_percentage': 0.0,
            'infra_license_cost': 0.0,
            'other_costs': 0.0,
            'monthly_savings': 0.0,
            'annual_savings': 0.0,
            'roi_first_year': 0.0,
            'roi_percentage_first_year': 0.0,
            'payback_period_months': 0.0,
        }
        
        pdf_success, pdf_buffer, pdf_error = ExportManager.export_to_pdf([zero_calc])
        excel_success, excel_buffer, excel_error = ExportManager.export_to_excel([zero_calc])
        
        assert pdf_success is True
        assert excel_success is True
        assert pdf_buffer is not None
        assert excel_buffer is not None
